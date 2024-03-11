from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.properties import StringProperty, ObjectProperty
from kivy.core.window import Window
from openpyxl import Workbook, load_workbook


#Builder.load_file("registrationsys.kv")
Window.clearcolor = "#79A3D2"#(23/255.0, 69/255.0, 23/255.0, 1)
class Container(BoxLayout):
	licenseNumber = ObjectProperty(None)
	name = ObjectProperty(None)
	surname = ObjectProperty(None)
	dob = ObjectProperty(None)
	nationality = ObjectProperty(None)
	club = ObjectProperty(None)
	position = ObjectProperty(None)
	debut = ObjectProperty(None)

	def saveit(self):
		
		wb = load_workbook("leagueDB.xlsx")
		ws = wb.active
		ws.append([self.name.text, self.surname.text, self.dob.text, self.nationality.text, self.club.text, self.position.text, self.debut.text, self.licenseNumber.text])
		wb.save("leagueDB.xlsx")
		#print("this is {} {} \n born: {} \n nationality: {}".format( self.name.text, self.surname.text, self.dob.text, self.nationality.text))
		#print("Plays for: {}\n Position:{}\n debuted:{}".format(self.club.text, self.position.text, self.debut.text))
		self.name.text = ""
		self.surname.text = ""
		self.dob.text = ""
		self.nationality.text = ""
		self.club.text = ""
		self.position.text = ""
		self.debut.text = ""
		self.licenseNumber.text = ""
		print("Successfully ran")
	def searchit(self):
		wb = load_workbook("leagueDB.xlsx")
		ws = wb.active
		searchstring = ObjectProperty(None)
		
		# This is the actual code for the search feature
		for row in range(2, ws.max_row + 1):
			for column in "ABCDEFGH":
				cell_name = "{}{}".format(column, row)
				if ws[cell_name].value == self.searchstring.text:

					'''
						28-06-23
						Update to make instead of searching for anything in the players credetials just limit the search to
						his licende number since it can not be duplicated
						the code should look like this:
							for row in range(2, ws.max_row+1):
								cell_name = "{}{}".format(license column, row)
								if ws[cell_name].value == self.searchstring.text:
					'''
					
					print("Successfully ran")
					self.searchstring.text = ""

					# This is the part that display this result of the research 
					for col in "ABCDEFGH":
						cell = "{}{}".format(col,row)
						xk = ws[cell].value
						if col == "A":
							self.name.text = xk
						elif col == "B":
							self.surname.text = xk
						elif col == "C":
							self.dob.text = xk
						elif col == "D":
							self.nationality.text = xk
						elif col == "E":
							self.club.text = xk
						elif col == "F":
							self.position.text = xk
						elif col == "G":
							self.debut.text = xk
						elif col == "H":
							self.licenseNumber.text = xk

						#print(xk)
		#print("yay")

	def updateit(self):
		wb = load_workbook("leagueDB.xlsx")
		ws = wb.active
		ws.append([self.name.text, self.surname.text, self.dob.text, self.nationality.text, self.club.text, self.position.text, self.debut.text, self.licenseNumber.text])
		wb.save("leagueDB.xlsx")
		
		self.name.text = ""
		self.surname.text = ""
		self.dob.text = ""
		self.nationality.text = ""
		self.club.text = ""
		self.position.text = ""
		self.debut.text = ""
		self.licenseNumber.text = ""

	def deleteit(self):
		pass


class SystemApp(App):
	pass


SystemApp().run()